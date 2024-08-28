# orders/views.py
import openpyxl
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
import io
import json
import os

# Path to your Excel file
EXCEL_FILE_PATH = '/home/mahmoud/orders.xlsx'

def read_excel_file():
    if not os.path.exists(EXCEL_FILE_PATH):
        return []

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)
        worksheet = workbook.active
        data = []

        # Read headers
        headers = [cell.value for cell in worksheet[1]]

        # Read rows
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)

        return data
    except Exception as e:
        raise RuntimeError(f"Error reading Excel file: {str(e)}")

def write_to_excel(data):
    headers = ['orderID', 'orderName', 'userName', 'status']
    
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Orders"
            worksheet.append(headers)
            workbook.save(EXCEL_FILE_PATH)
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        
        worksheet = workbook.active

        for item in data:
            if all(field in item for field in headers):
                worksheet.append([item[field] for field in headers])
            else:
                raise ValueError("Invalid data format")

        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        with open(EXCEL_FILE_PATH, 'wb') as file:
            file.write(file_stream.getvalue())

    except Exception as e:
        raise RuntimeError(f"Error writing to Excel file: {str(e)}")

@csrf_exempt
def order_view(request):
    if request.method == 'GET':
        try:
            data = read_excel_file()
            return render(request, 'home.html', {'data': data})
        except RuntimeError as e:
            return HttpResponseBadRequest(str(e))
        
    elif request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))

            if not isinstance(data, list) or not all(isinstance(item, dict) for item in data):
                return HttpResponseBadRequest("Invalid JSON format. Expected a list of dictionaries.")

            write_to_excel(data)
            print(data)
            return JsonResponse({'message': 'Excel file updated successfully'})

        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON data")
        except ValueError as e:
            return HttpResponseBadRequest(str(e))
        except Exception as e:
            return HttpResponseBadRequest(f"Error processing request: {str(e)}")
